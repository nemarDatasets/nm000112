"""Finer-grained Affective Computing EEG Dataset (FACED)

Download command:
```bash
https://www.synapse.org/#!Synapse:syn50614194
```
"""

README_CONTENT = """**Introduction:**
The Finer-grained Affective Computing EEG Dataset (FACED) contains scalp EEG recordings from 123 healthy participants who watched 28 emotion-eliciting video clips designed to evoke nine different emotion categories. The dataset includes four negative emotions (anger, fear, disgust, sadness) from Ekman's basic emotions and four positive emotions (amusement, inspiration, joy, tenderness) selected based on recent psychological and neuroscience progress and application needs. Participants provided detailed self-reported emotion ratings on 12 dimensions: eight emotions, arousal, valence, liking, and familiarity. The dataset is designed to facilitate cross-subject affective computing research and development of EEG-based emotion recognition algorithms for real-world applications.

**Overview of the experiment:**
Participants (123 subjects, 75 female, ages 17-38, mean=23.2 years) were seated 60 cm from a 22-inch LCD monitor in a regular office environment. Each trial consisted of: (1) a 5-second fixation cross, (2) a video clip of varying length (typically 30-60 seconds), and (3) subjective emotional rating on 12 items (anger, fear, disgust, sadness, amusement, inspiration, joy, tenderness, valence, arousal, liking, familiarity) on a continuous 0-7 scale, followed by at least 30 seconds rest. Video clips were presented in blocks: three positive blocks, three negative blocks, and one neutral block, with 20 arithmetic problems between blocks to minimize carryover effects. The 28 video clips were designed to target nine emotion categories, with randomized presentation order across participants. EEG was recorded using a 32-channel biosignal recording system sampled at either 1000 Hz (92 subjects) or 250 Hz (31 subjects), with channels positioned according to the International 10-20 system. Signal units were recorded in either Volts or microVolts depending on the hardware configuration used.

**Video stimulus information:**
The dataset includes 28 video clips designed to elicit nine emotion categories (Trigger values 1–28):
- Anger (Videos 1-3): Durations 73-81 seconds, negative valence
- Disgust (Videos 4-6): Durations 69-91 seconds, negative valence
- Fear (Videos 7-9): Durations 56-106 seconds, negative valence
- Sadness (Videos 10-12): Durations 45-82 seconds, negative valence
- Neutral (Videos 13-16): Durations 35-43 seconds, neutral valence
- Amusement (Videos 17-19): Durations 56-73 seconds, positive valence
- Inspiration (Videos 20-22): Durations 76-129 seconds, positive valence
- Joy (Videos 23-25): Durations 34-68 seconds, positive valence
- Tenderness (Videos 26-28): Durations 54-77 seconds, positive valence

Metadata for each video (duration, source film, source database, valence, targeted emotion) is read from Stimuli_info.xlsx.

**Event markers (from evt.bdf annotations):**
- 100: Task/block start
- 101: Video onset
- 102: Video offset
- 1–28: Video index (appears just before 101, used to link to stimulus metadata)
- 201/202: Block boundary markers
- "Start Impedance" / "Stop Impedance": Technical markers (ignored)

The conversion script reads evt.bdf annotations for each subject, parses video presentation spans (from video index + 101 to 102), and creates MNE Annotations with the source film title (video_title) as description. These annotations are exported to BIDS events.tsv with extra columns:
- emotion_label: targeted emotion category (Anger, Disgust, Fear, Sadness, Neutral, Amusement, Inspiration, Joy, Tenderness)
- binary_label: positive/negative/neutral classification
- video_index: 1–28
- Self-reported ratings (Joy, Tenderness, Inspiration, Amusement, Anger, Disgust, Fear, Sadness, Arousal, Valence, Familiarity, Liking)

**Description of the preprocessing if any:**
Raw BDF files from the biosignal recording system have been converted to BIDS format. Channel names are standardized to match the International 10-20 nomenclature. Subjects have been assigned numeric IDs (sub-000 through sub-122) corresponding to their original subject designations in the dataset. Recording dates have been set to a default value (2023-01-01) due to privacy considerations, while time relationships between files are preserved. Subject demographic information (age, sex) has been extracted from the Recording_info.csv file and properly formatted for BIDS.

Stimulus timing information from the evt.bdf event files has been parsed and enriched with metadata from Stimuli_info.xlsx. Each video presentation is annotated with the targeted emotion category (Anger, Disgust, Fear, Sadness, Neutral, Amusement, Inspiration, Joy, Tenderness) and includes self-reported ratings from After_remarks.mat when available.

**Citation:**
When using this dataset, please cite:

1. Liu, Y., Sourina, O., & Nguyen, M. K. (2023). Finer-grained Affective Computing EEG Dataset. Scientific Data, 10(1), 809. https://doi.org/10.1038/s41597-023-02650-w

2. Synapse Platform: https://www.synapse.org/#!Synapse:syn50614194

3. The dataset is available at the Synapse platform repository.

**Data curators:**
Pierre Guetschel (BIDS conversion)

Original data collection team:
- Yisi Liu (Nanyang Technological University)
- Olga Sourina (Nanyang Technological University)
- Minh Khoa Nguyen (Nanyang Technological University)
"""

DATASET_NAME = "FACED - Finer-grained Affective Computing EEG Dataset"

from pathlib import Path
import shutil
import datetime
import warnings

import pandas as pd
import numpy as np
from scipy.io import loadmat
from mne.io import read_raw_bdf
from mne import read_annotations, Annotations
from mne_bids import BIDSPath, write_raw_bids, make_dataset_description, make_report
from openpyxl import load_workbook


def _load_stimuli_info(source_root: Path) -> pd.DataFrame:
    """Load stimulus metadata from Stimuli_info.xlsx.

    Returns DataFrame with columns: video_index, duration, source_film,
    source_database, valence, targeted_emotion.
    """
    stimuli_path = source_root / "Stimuli_info.xlsx"
    if not stimuli_path.exists():
        raise FileNotFoundError(f"Stimuli_info.xlsx not found at {stimuli_path}")

    wb = load_workbook(stimuli_path)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)
    data = []
    for row in rows:
        if row[0] is None:
            break
        data.append(
            {
                "video_index": int(row[0]),
                "duration": float(row[1]) if row[1] is not None else None,
                "source_film": row[2],
                "source_database": row[3],
                "valence": row[4],
                "targeted_emotion": row[5],
            }
        )
    return pd.DataFrame(data)


def _load_task_event_map(source_root: Path) -> tuple[dict, set]:
    """Parse Task_event.xlsx to validate and return trigger mappings.

    Returns (event_map, valid_videos) where event_map maps integer codes to
    canonical labels (task_block_start, video_start, video_end) and
    valid_videos is the set of video index codes discovered (expected 1..28).
    """
    task_path = source_root / "Task_event.xlsx"
    if not task_path.exists():
        raise FileNotFoundError(f"Task_event.xlsx not found at {task_path}")

    wb = load_workbook(task_path)
    ws = wb.active

    found_codes = set()
    valid_videos: set[int] = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                val = int(cell)
                if val in (100, 101, 102):
                    found_codes.add(val)
                if 1 <= val <= 28:
                    valid_videos.add(val)

    required = {100, 101, 102}
    if not required.issubset(found_codes):
        missing = sorted(list(required - found_codes))
        raise ValueError(f"Task_event.xlsx missing required trigger codes: {missing}")

    # Require full set of 1..28 video indices
    expected_videos = set(range(1, 29))
    if not expected_videos.issubset(valid_videos):
        missing_videos = sorted(list(expected_videos - valid_videos))
        raise ValueError(f"Task_event.xlsx missing video indices: {missing_videos}")

    event_map = {100: "task_block_start", 101: "video_start", 102: "video_end"}
    return event_map, valid_videos


def _validate_video_index_alignment(source_root: Path) -> set:
    """Ensure Stimuli_info.xlsx has a complete, unique set of video_index 1..28
    and that it matches the indices discovered in Task_event.xlsx.

    Returns the set of video indices if valid, else raises ValueError.
    """
    stimuli_df = _load_stimuli_info(source_root)

    if "video_index" not in stimuli_df.columns:
        raise ValueError("Stimuli_info.xlsx missing required 'video_index' column")

    # Normalize to integers and validate
    try:
        indices = stimuli_df["video_index"].astype(int).tolist()
    except Exception as e:
        raise ValueError(f"Invalid video_index values in Stimuli_info.xlsx: {e}")

    index_set = set(indices)
    expected = set(range(1, 29))

    # Check for duplicates
    if len(indices) != len(index_set):
        # Find duplicate values
        seen = set()
        duplicates = set()
        for i in indices:
            if i in seen:
                duplicates.add(i)
            seen.add(i)
        dup_str = ", ".join(str(d) for d in sorted(duplicates))
        raise ValueError(
            f"Stimuli_info.xlsx has duplicate video_index values: {dup_str}"
        )

    # Check for completeness 1..28
    if index_set != expected:
        missing = sorted(list(expected - index_set))
        extra = sorted(list(index_set - expected))
        parts = []
        if missing:
            parts.append(f"missing {missing}")
        if extra:
            parts.append(f"unexpected {extra}")
        detail = "; ".join(parts) if parts else "mismatch"
        raise ValueError(
            f"Stimuli_info.xlsx video_index set does not equal 1..28: {detail}"
        )

    # Cross-check with Task_event.xlsx discovery
    _, valid_videos = _load_task_event_map(source_root)
    if index_set != valid_videos:
        missing_in_task = sorted(list(index_set - valid_videos))
        missing_in_stim = sorted(list(valid_videos - index_set))
        raise ValueError(
            "Stimuli_info.xlsx and Task_event.xlsx have mismatched video indices: "
            f"only-in-stimuli={missing_in_task}, only-in-task={missing_in_stim}"
        )

    return index_set


def _load_video_metadata_map(source_root: Path) -> dict:
    """Build a mapping from video_index to metadata from Stimuli_info.xlsx.

    Returns dict[int, dict] with keys: emotion, valence, title, binary_label.
    """
    stimuli_df = _load_stimuli_info(source_root)
    result = {}
    for _, row in stimuli_df.iterrows():
        idx = int(row.video_index)
        valence_str = str(row.valence or "").lower()
        if valence_str.startswith("pos"):
            binary_label = "positive"
        elif valence_str.startswith("neg"):
            binary_label = "negative"
        else:
            binary_label = "neutral"
        result[idx] = {
            "emotion": row.targeted_emotion,
            "valence": row.valence,
            "title": row.source_film,
            "binary_label": binary_label,
        }
    return result


# Order of the ratings, taken from Code/Readme.md :
RATING_KEYS = [
    "Joy",
    "Tenderness",
    "Inspiration",
    "Amusement",
    "Anger",
    "Disgust",
    "Fear",
    "Sadness",
    "Arousal",
    "Valence",
    "Familiarity",
    "Liking",
]


def _load_all_ratings(source_root: Path, subject_id: str) -> dict:
    """Load self-reported ratings for all videos from After_remarks.mat.

    The After_remarks.mat file contains a struct array with 28 entries (one per trial).
    Each entry has:
    - score: array of 12 rating values in order matching RATING_KEYS
    - trial: trial number (1-28)
    - vid: video index (1-28)
    - Accuracy, ResponseTime: additional fields (unused)

    Returns a dict mapping video_index (1-28) to a dict of rating values.
    """
    mat_path = source_root / "Data" / f"sub{subject_id}" / "After_remarks.mat"
    if not mat_path.exists():
        raise FileNotFoundError(f"After_remarks.mat not found for subject {subject_id}")

    data = loadmat(mat_path, squeeze_me=True, struct_as_record=False)
    arr = data["After_remark"]

    all_ratings = {}
    for item in arr:
        try:
            video_idx = int(getattr(item, "vid", 0))
            if not (1 <= video_idx <= 28):
                continue
            score = getattr(item, "score", None)
            if score is None or not hasattr(score, "__len__"):
                continue
            # score is an array of 12 values matching RATING_KEYS order
            ratings = {}
            for i, key in enumerate(RATING_KEYS):
                if i < len(score):
                    val = score[i]
                    if isinstance(val, np.ndarray):
                        val = val.item() if val.size == 1 else None
                    ratings[key] = float(val) if val is not None else None
                else:
                    ratings[key] = None
            all_ratings[video_idx] = ratings
        except Exception:
            continue

    return all_ratings


def _create_annotations_from_evt(
    raw, evt_path: Path, source_root: Path, subject_id: str, video_meta: dict
):
    """Create annotations from evt.bdf events file.

    Parses the evt.bdf annotations to extract video presentation spans:
    - video_index (1-28) appears just before 101 (video start)
    - 102 marks video end
    - Each span is annotated with the video title as description

    Sets raw.annotations with extras for BIDS events.tsv including:
    onset, duration, trial_type (video_title), emotion_label, binary_label, video_index, and 12 rating columns.
    """
    # Load ratings for this subject
    all_ratings = _load_all_ratings(source_root, subject_id)

    # Read events from evt.bdf
    evt_annot = read_annotations(str(evt_path))
    events = list(zip(evt_annot.onset.tolist(), evt_annot.description.tolist()))

    # Parse events to extract video spans
    spans = []  # list of (onset, duration, video_index)
    current_video_idx = None
    current_start = None

    for onset, desc in events:
        code = desc.strip()

        # Check if this is a video index (1-28)
        if not code.isdigit():  # keep other codes as is
            spans.append((onset, 0.0, code))
            continue
        code = int(code)
        if 1 <= code <= 28:
            current_video_idx = code
        elif code == 101:
            # Video start
            current_start = onset
        elif code == 102:
            # Video end
            if current_start is None or current_video_idx is None:
                raise ValueError(
                    f"Malformed evt.bdf: video end (102) without start for subject {subject_id}"
                )
            duration = onset - current_start
            spans.append((current_start, duration, current_video_idx))
            current_start = None
            current_video_idx = None
        elif code == 100:
            spans.append((onset, 0.0, "Experiment start"))
        else:  # keep other codes as is
            spans.append((onset, 0.0, code))

    # Build annotations with extras for BIDS export
    onsets = []
    durations = []
    descriptions = []
    extras = []

    for span_onset, span_duration, code in spans:
        onsets.append(span_onset)
        durations.append(span_duration)

        meta = video_meta.get(code, None)
        if meta is None:
            descriptions.append(code)
            extras.append({})
            continue

        descriptions.append(meta["title"])

        # Build extras dict for this annotation
        extra = {
            "emotion_label": meta["emotion"],
            "binary_label": meta["binary_label"],
            "video_index": code,
        }
        # Add self-ratings
        ratings = all_ratings[code]
        for key in RATING_KEYS:
            extra[key] = ratings.get(key)
        extras.append(extra)

    raw.set_annotations(Annotations(onsets, durations, descriptions, extras=extras))


def _get_records(source_root: Path):
    """Extract records from FACED dataset structure.

    The FACED dataset organizes each subject in a folder: Data/sub000/, Data/sub001/, etc.
    Each subject folder contains:
    - data.bdf: the main EEG recording
    - evt.bdf: event/trigger information with video onset/offset markers

    Yields tuples of (data_bdf_path, evt_bdf_path, bids_path).
    """
    data_root = source_root / "Data"

    # Find all subject folders
    subject_folders = sorted([d for d in data_root.iterdir() if d.is_dir()])

    for subject_folder in subject_folders:
        # Extract subject ID (e.g., "sub000" -> "000")
        subject_id = subject_folder.name
        if not subject_id.startswith("sub"):
            continue

        bdf_file = subject_folder / "data.bdf"
        evt_file = subject_folder / "evt.bdf"
        if not bdf_file.exists():
            warnings.warn(f"Missing data.bdf in {subject_folder}")
            continue
        if not evt_file.exists():
            warnings.warn(f"Missing evt.bdf in {subject_folder}")
            continue

        # Create BIDS path for this recording
        bids_path = BIDSPath(
            subject=subject_id.replace("sub", ""),
            task="watchingVideoClips",
            suffix="eeg",
            datatype="eeg",
            extension=".bdf",
        )

        yield str(bdf_file), str(evt_file), bids_path


def main(
    source_root: Path,
    bids_root: Path,
    overwrite: bool = False,
    validate_only: bool = False,
    subject: str | None = None,
    finalize_only: bool = False,
):
    """Convert the FACED dataset to BIDS format.

    Parameters
    ----------
    source_root : Path
        Path to the root folder of the FACED dataset.
        Downloaded from https://www.synapse.org/#!Synapse:syn50614194
    bids_root : Path
        Path to the root of the BIDS dataset to create.
    overwrite : bool
        If True, overwrite existing BIDS files.
    validate_only : bool
        If True, only validate Stimuli_info.xlsx / Task_event.xlsx alignment.
    subject : str | None
        Optional subject ID to convert (e.g., "000" or "sub000"). If provided,
        only this subject is processed. Useful for testing.
    """
    source_root = Path(source_root).expanduser()
    bids_root = Path(bids_root).expanduser()

    # Validate Stimuli_info.xlsx and Task_event.xlsx alignment before processing
    aligned_indices = _validate_video_index_alignment(source_root)
    print(f"Validation successful. video_index set: {sorted(aligned_indices)}")
    if validate_only:
        return

    records = list(_get_records(source_root))

    # Filter to single subject if specified
    if subject is not None:
        # Normalize subject ID (remove "sub" prefix if present, handle int from CLI)
        subj_id = str(subject).replace("sub", "").zfill(3)
        records = [(d, e, b) for d, e, b in records if b.subject == subj_id]
        if not records:
            raise ValueError(f"Subject '{subject}' not found in {source_root / 'Data'}")
        print(f"Processing single subject: sub{subj_id}")

    if not records:
        raise ValueError(f"No EEG records found in {source_root / 'Data'}")

    # Read subject information from Recording_info.csv
    recording_info_path = source_root / "Recording_info.csv"
    recording_info = pd.read_csv(recording_info_path)

    # Create mapping from original subject identifier to recording info
    # The CSV uses 'sub' column with values like 'sub000', 'sub001', etc.
    subject_info_map = {}
    for _, row in recording_info.iterrows():
        subject_id = row["sub"].replace("sub", "")
        subject_info_map[subject_id] = {
            "gender": row["Gender"],
            "age": row["Age"],
            "sample_rate": row["Sample_rate"],
            "unit": row["Unit"],
        }

    # Create BIDS root directory
    bids_root.mkdir(parents=True, exist_ok=True)

    # Update all bids paths with root
    for _, _, bids_path in records:
        bids_path = bids_path.update(root=bids_root)

    if finalize_only:
        _finalize_dataset(bids_root, overwrite=overwrite)
        return

    # Sanity check: no duplicate BIDS paths
    bids_paths = [bids_path.fpath for _, _, bids_path in records]
    assert len(bids_paths) == len(set(bids_paths)), "Duplicate BIDS paths found"

    # Load video metadata once for all subjects
    video_meta = _load_video_metadata_map(source_root)

    # Process each record
    for eeg_file_path, evt_file_path, bids_path in records:
        subject_id = bids_path.subject
        if not overwrite and bids_path.fpath.exists():
            print(f"Skipping {bids_path.fpath} (already exists)")
            continue

        # Read EEG data
        raw = read_raw_bdf(eeg_file_path, preload=False, verbose=False)

        # Create annotations from evt.bdf events (sets raw.annotations with extras)
        _create_annotations_from_evt(
            raw, Path(evt_file_path), source_root, subject_id, video_meta
        )

        # Get subject info
        if subject_id in subject_info_map:
            info = subject_info_map[subject_id]

            # Set subject info
            gender_map = {"M": 1, "F": 2}
            sex = gender_map.get(info["gender"], 0)

            # Calculate birth date from age and recording date
            # Use a default recording date of 2023-01-01 for privacy
            recording_date = datetime.date(2023, 1, 1)
            birth_date = recording_date - datetime.timedelta(days=info["age"] * 365.25)

            raw.info["subject_info"] = {
                "his_id": f"sub{subject_id}",
                "birthday": birth_date,
                "sex": int(sex),
            }
            raw.set_meas_date(
                datetime.datetime(2023, 1, 1, tzinfo=datetime.timezone.utc)
            )

        # Note: Unit conversion handled appropriately by MNE-BIDS based on the data units
        # Some recordings in the dataset are in V, others in uV - MNE handles this properly

        # Write to BIDS - annotations are automatically exported to events.tsv
        write_raw_bids(
            raw,
            bids_path,
            overwrite=True,
            verbose=False,
        )
        print(f"Converted: {bids_path.fpath}")

    _finalize_dataset(bids_root, overwrite=overwrite)


def _finalize_dataset(bids_root: Path, overwrite: bool = False):
    """Finalize BIDS dataset: save script, dataset_description, report, README, cleanup."""
    # Save conversion script
    script_path = Path(__file__)
    script_dest = bids_root / "code" / script_path.name
    script_dest.parent.mkdir(exist_ok=True)
    shutil.copy2(script_path, script_dest)

    # Create dataset description
    description_file = bids_root / "dataset_description.json"
    if description_file.exists() and overwrite:
        description_file.unlink()
    make_dataset_description(
        path=bids_root,
        name=DATASET_NAME,
        dataset_type="derivative",
        references_and_links=[
            "https://doi.org/10.1038/s41597-023-02650-w",
            "https://www.synapse.org/#!Synapse:syn50614194",
        ],
        source_datasets=[
            {"DOI": "https://doi.org/10.1038/s41597-023-02650-w"},
        ],
        authors=["Pierre Guetschel"],
        overwrite=overwrite,
    )

    # Remove macOS resource fork files that can break make_report
    for dotfile in bids_root.rglob("._*"):
        dotfile.unlink()

    try:
        report_str = make_report(bids_root)
        print(report_str)
    except Exception as e:
        warnings.warn(f"make_report failed: {e}")
        report_str = str(e)

    # Write README (after make_dataset_description to ensure it's not overwritten)
    readme_path = bids_root / "README.md"
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(
            f"# {DATASET_NAME}\n\n{README_CONTENT}\n\n---\n\n"
            f"**Automatic report:**\n\n*Report automatically generated by `mne_bids.make_report()`.*\n\n> {report_str}"
        )

    # Remove participants.json if it exists
    participants_json = bids_root / "participants.json"
    if participants_json.exists():
        participants_json.unlink()
        print(f"Removed {participants_json}")

    # Clean up participants.tsv by removing columns where all values are empty or "n/a"
    participants_tsv = bids_root / "participants.tsv"
    if participants_tsv.exists():
        df = pd.read_csv(participants_tsv, sep="\t")
        # Find columns where all non-participant_id values are empty or "n/a"
        cols_to_drop = []
        for col in df.columns:
            if col != "participant_id":
                is_empty = df[col].isna() | (df[col] == "n/a")
                if is_empty.all():
                    cols_to_drop.append(col)
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)
            df.to_csv(participants_tsv, sep="\t", index=False)
            print(f"Removed empty columns from {participants_tsv}: {cols_to_drop}")


if __name__ == "__main__":
    from fire import Fire

    Fire(main)
    # python bids_maker/datasets/faced.py --source_root ~/data/FACED/ --bids_root ~/data/bids/faced/ --overwrite=True
    # python bids_maker/datasets/faced.py --source_root ~/data/FACED/ --bids_root ~/data/bids/faced/ --validate_only=True
    # python bids_maker/datasets/faced.py --source_root ~/data/FACED/ --bids_root ~/data/bids/faced/ --subject=000 --overwrite=True
